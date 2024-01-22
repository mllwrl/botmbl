#Imports
from difflib import SequenceMatcher
import telebot
from telebot import types
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
#token
token='***'
bot=telebot.TeleBot(token)
ChatID = None
#Excel
wb1 = load_workbook('Raspisanie_urokov_5_11_S_9_yanvarya_2024_KABINETY.xlsx')
sheet_ranges1 = wb1.active
#Mass
Query = ['Педагогический состав', 'Расписание', 'Экзамены', 'Поступление', 'Портфолио',  'Кружки', 'Подготовка', 'Подготовительные группы', 'Педагоги', 'Учителя']
QueryTeachers = ['Сыромолотова Татьяна Яковлевна','Асташкина Светлана Геннадьевна','Заносиенко Елена Владимировна','Сердюкова Лариса Валентиновна','Новичкова Татьяна Владимировна','Погодина Ольга Владимировна','Байдова Светлана Юрьевна','Егорова Елена Михайловна','Курашова Маргарита Александровна','Митузова Мария Павловна','Вачинич Наталья Николаевна','Швырева Елена Валерьевна','Бурова Светлана Григорьевна','Мальчикова Снежана Александровна','Исаева Любовь Геннадьевна','Балтаева Алина Сергеевна','Лункина Елена Викторовна','Васильева Людмила Леонидовна','Степанова Наталья Викторовна','Гордеева Марина Николаевна','Шилина Мария Сергеевна','Кондратьева Виктория Андреевна','Овчинникова Виктория Викторовна','Маркушина Татьяна Анатольевна','Бутримова Кристина Александровна','Осипова Любовь Владимировна','Костянчук Людмила Антоновна','Широких Ирина Геннадиевна','Голубев Сергей Владимирович','Голубева Ольга Игоревна','Чеблукова Екатерина Александровна','Сюсюкина Елена Николаевна','Сюсюкин Алексей Юрьевич','Володина Екатерина Владимировна','Жутаева Ольга Михайловна','Потехина Ирина Владимировна','Лашева Светлана Евгеньевна','Майорова Надежда Владимировна','Подкидышева Лилия Валерьевна','Ерилова Екатерина Александровна']
vivod = ''
mesID = None
msgTP = None
testVivod = False
#NotTGFunction
def similarity_score(query1, query2):
    similarity = SequenceMatcher(None, query1, query2).ratio()
    return similarity

#TGcommands
@bot.message_handler(commands=['start', 'help', 'continue'])
def start_message(message):
    global ChatID
    ChatID = message.chat.id
    if message.text=="/start":
        ChatID = message.chat.id
        markup=types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1=types.KeyboardButton("Задать вопрос")
        item2=types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id,"Привет. Если бот не реагирует на ваши действия, воспользуйтесь командой /continue или /help", reply_markup=markup)
    if message.text == "/help":
        bot.send_message(message.chat.id, "1. Если Вы хотите задать вопрос, то воспользуйтесь кнопкой 'Задать вопрос'\n2. Если Вы обнаружили неисправность, то сообщить об этом можно, нажав на кнопку 'Сообщить о неполадках'\n3. Хотите добавить готовые ответы или предложить новую идею? Воспользуйтесь кнопкой 'Предложить идею'\n4. Если Вам не отвечает тех. поддержка, воспользуйтесь кнопкой 'Сообщить о неполадках', опишите вашу проблему и запустите бота @mblansbot\n5. Если вы считаете, что в помощи нужна дополнительная информация, воспользуйтесь кнопкой 'Предложить идею'\n7. Желательно, указать имя пользователя в Вашем аккаунте (имя, начинающееся с @)")
    if message.text == "/continue":
        ChatID = message.chat.id
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id, "Выполнено", reply_markup=markup)
@bot.message_handler(content_types='text')
def message_reply(message):
    global ChatID, mesID
    ChatID = message.chat.id
    if message.text == "Задать вопрос":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Назад")
        markup.add(item1)
        PrMessage = "Задайте ваш вопрос:"
        msg = bot.send_message(message.chat.id, PrMessage, reply_markup=markup)
        bot.register_next_step_handler(msg, setPR)
    if message.text == "Сообщить о неполадках":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Назад")
        markup.add(item1)
        PrMessage = "Что нужно сообщить:"
        msg = bot.send_message(message.chat.id, PrMessage, reply_markup=markup)
        bot.register_next_step_handler(msg, nep)
    if message.text == "Предложить идею":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Назад")
        markup.add(item1)
        PrMessage = "Что нужно сообщить:"
        msg = bot.send_message(message.chat.id, PrMessage, reply_markup=markup)
        bot.register_next_step_handler(msg, idea)
    if message.text == 'Быстрый доступ':
        markup = types.InlineKeyboardMarkup()
        item1 = types.InlineKeyboardButton(text='Список педагогического состава', callback_data='список')
        item2 = types.InlineKeyboardButton(text='Найти учителя по имени', callback_data='НПИ')
        item3 = types.InlineKeyboardButton(text='Ученическое расписание', callback_data='УченРасп')
        item4 = types.InlineKeyboardButton(text='Расписание учителей', callback_data='УчитРасп')
        item5 = types.InlineKeyboardButton(text='Поступление', callback_data='Поступ')
        markup.add(item1, item2, item3, item4, item5)
        mesID = bot.send_message(ChatID, 'Быстрый доступ:', reply_markup=markup)

def idea(message):
    if message.text=="Назад":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id, "Действие отменено", reply_markup=markup)
        return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Задать вопрос")
    item2 = types.KeyboardButton("Сообщить о неполадках")
    item3 = types.KeyboardButton("Предложить идею")
    item4 = types.KeyboardButton("Быстрый доступ")
    markup.add(item1, item2, item3, item4)
    bot.send_message(message.chat.id, "Отправлено!", reply_markup=markup)
    bot.send_message('-1001826121215', f"@mllwrl $IDEA    $DEBUGUSER (CHATID:{message.from_user.id}, UNAME: @{message.from_user.username}, NAME: {message.from_user.first_name} {message.from_user.last_name}), MESSAGE: {message.text}", reply_markup=markup)

def nep(message):
    if message.text=="Назад":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(message.chat.id, "Действие отменено", reply_markup=markup)
        return
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Задать вопрос")
    item2 = types.KeyboardButton("Сообщить о неполадках")
    item3 = types.KeyboardButton("Предложить идею")
    item4 = types.KeyboardButton("Быстрый доступ")
    markup.add(item1, item2, item3, item4)
    bot.send_message(message.chat.id, "Отправлено!", reply_markup=markup)
    bot.send_message('-1001826121215', f"@mllwrl $NEPOLADKA    $DEBUGUSER (CHATID:{message.from_user.id}, UNAME: @{message.from_user.username}, NAME: {message.from_user.first_name} {message.from_user.last_name}), MESSAGE: {message.text}", reply_markup=markup)

def setPR(message):
    if message.text=="Назад":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        markup.add(item1, item2, item3)
        bot.send_message(message.chat.id, "Действие отменено", reply_markup=markup)
        return
    maxes = 0
    global vivod, mesID, testVivod
    vivod = ''
    markup = types.InlineKeyboardMarkup()
    for i in Query:
        score = similarity_score(i, message.text)
        if maxes < score:
            maxes = score
            vivod = i
    print(f"$DEBUGUSER (CHATID:{message.from_user.id}, UNAME: @{message.from_user.username}, NAME: {message.from_user.first_name} {message.from_user.last_name})")
    bot.send_message('-1001826121215', f"$DEBUGUSER (CHATID:{message.from_user.id}, UNAME: @{message.from_user.username}, NAME: {message.from_user.first_name} {message.from_user.last_name}), MESSAGE: {message.text}", reply_markup=markup)
    if vivod == '':
        testVivod = False
        item3 = types.InlineKeyboardButton(text='Да', callback_data='Да2')
        item4 = types.InlineKeyboardButton(text='Нет', callback_data='Нет2')
        markup.add(item3, item4)
        bot.send_message(ChatID, 'Нет ответа на ваш вопрос, обратиться к тех. поддержке?', reply_markup=markup)
    else:
        item1 = types.InlineKeyboardButton(text='Да', callback_data='Да1')
        item2 = types.InlineKeyboardButton(text='Нет', callback_data='Нет1')
        markup.add(item1, item2)
        mesID = bot.send_message(message.chat.id, 'Вы имели ввиду это:'+'\n'+vivod+'?', reply_markup=markup)

@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    global vivod, ChatID, msgTP, testVivod, mesID
    if call.data == "Да1":
        markup = types.InlineKeyboardMarkup()
        if vivod == 'Учителя' or vivod == 'Педагогический состав' or vivod == 'Педагоги':
            item1 = types.InlineKeyboardButton(text='Список состава', callback_data='список')
            item2 = types.InlineKeyboardButton(text='Найти по имени', callback_data='НПИ')
            markup.add(item1, item2)
            bot.edit_message_text(text='Вывести список или найти учителя по имени?', message_id=mesID.id, chat_id=ChatID, reply_markup=markup)
        if vivod == 'Расписание':
            item1 = types.InlineKeyboardButton(text='Для учеников', callback_data='УченРасп')
            item2 = types.InlineKeyboardButton(text='Для учителей', callback_data='УчитРасп')
            markup.add(item1, item2)
            bot.edit_message_text(text='Какое из расписания вас интересует?', message_id=mesID.id,chat_id=ChatID, reply_markup=markup)
        if vivod == 'Поступление':
            item1 = types.InlineKeyboardButton(text='Поступление в 1 класс', url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/buduschim-pervoklassnikam/')
            item2 = types.InlineKeyboardButton(text='Правила поступления', url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/pravila-priema-perevoda-otchisleniya/')
            item3 = types.InlineKeyboardButton(text='Поступление в 10 класс',url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/priem-v-10-klass/')
            item4 = types.InlineKeyboardButton(text='Образец заявления в 1 класс', url='https://docs.google.com/document/d/1AIQ1sFvuAB833WpN-NoFUnYaNlb7zh95/edit?usp=share_link&ouid=114487781857733996831&rtpof=true&sd=true')
            item5 = types.InlineKeyboardButton(text='Закрепленные территории', callback_data='ЗакрТер')
            item6 = types.InlineKeyboardButton(text='Перечень документов в 1 класс', callback_data='Доки1к')
            markup.add(item1, item2, item3, item4,item5, item6)
            bot.edit_message_text(text='Выберите нужный вариант:', message_id=mesID.id, chat_id=ChatID, reply_markup=markup)
    #Расписание
    if call.data == 'УченРасп':
        msg = bot.send_message(ChatID, text='Введите ваш класс')
        bot.register_next_step_handler(msg, RaspUCHEN)
    if call.data == 'УчитРасп':
        msg = bot.send_message(ChatID, text='Введите вашу фамилию и инициалы:')
        bot.register_next_step_handler(msg, RaspUCIT)
    #Поступление
    if call.data == 'Поступ':
        markup=types.InlineKeyboardMarkup()
        item1 = types.InlineKeyboardButton(text='Поступление в 1 класс',url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/buduschim-pervoklassnikam/')
        item2 = types.InlineKeyboardButton(text='Правила поступления',url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/pravila-priema-perevoda-otchisleniya/')
        item3 = types.InlineKeyboardButton(text='Поступление в 10 класс', url='https://mbl-sar.gosuslugi.ru/roditelyam-i-uchenikam/poleznaya-informatsiya/priem-v-10-klass/')
        item4 = types.InlineKeyboardButton(text='Образец заявления в 1 класс',url='https://docs.google.com/document/d/1AIQ1sFvuAB833WpN-NoFUnYaNlb7zh95/edit?usp=share_link&ouid=114487781857733996831&rtpof=true&sd=true')
        item5 = types.InlineKeyboardButton(text='Закрепленные территории', callback_data='ЗакрТер')
        item6 = types.InlineKeyboardButton(text='Перечень документов в 1 класс', callback_data='Доки1к')
        markup.add(item1, item2, item3, item4, item5, item6)
        bot.send_message( chat_id=ChatID, text='Выберите нужный вариант:', reply_markup=markup)
    if call.data == 'ЗакрТер':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(ChatID, text='1-й проезд им. Блинова Ф.А. все дома\n2-й проезд им. Блинова Ф.А. дома 2, 3, 3а, 3б, 4, 4а, 6, 6а, 4 б, 6 б, 6 в, 8, 8а, 9а\n3-й проезд им. Блинова Ф.А. дома 2, 4\nул. Романтиков дома 44, 46, 46а, 46б, 48, 48а,48б\nул. им. Академика О.К. Антонова дома 26, 26а, 26б, 26в, 26г, 24, 24а, 24б, 24в, 24г\nул. им. Блинова Ф.А. дома 25, 25а, 25б, 52, 52а, 52б, 52в, 19, 21, 21а, 21б\nул. Медовая дома 2, 5, 6, 7, 9, 12, 13, 15, 23, 25, 27, 28, 29, 31, 33, 35, 36, 39, 40, 41, 42, 43, 46, 50, 52, 55а, 56, 57, 57а, 59, 60\n1-й Медовый проезд дома 1, 2, 3, 5, 6, 7, 8а, 9, 12, 14, 18, 19, 27\n2-й Медовый проезд дома 8, 10, 16, 20, 24, 30, 32, 39, 41, 45, 47, 48, 50, 53, 55, 61, 62, 73, 75, 77, 84\n1-й Терновый проезд дома 3, 8\n2-й Терновый проезд дома 1, 4, 5, 7, 41\n3-й Терновый проезд дома 1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 17\n4-й Терновый проезд дома 1, 2, 3, 4, 5, 6, 9, 13, 15, 17, 18, 19, 21', reply_markup=markup)
    if call.data == 'Доки1к':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(ChatID, text='*Перечень документов, предоставляемых при поступлении в 1 класс*\nДля приема в 1 класс родители (законные представители) детей, предъявляют следующие документы:\n1)    копию документа, удостоверяющего личность родителя\n2)    копию свидетельства о рождении ребенка; \n3)    копию документа о регистрации ребенка по месту жительства или по месту пребывания на закрепленной территории или справку о приеме документов для оформления регистрации по месту жительства;\n4)    копию свидетельства о рождении полнородных и неполнородных брата и (или)сестры (в случае использования права преимущественного приема на обучение по образовательным программам начального общего образования)\n5)    справку с места работы родителя (при наличии права первоочередного приема на обучение);\n6)    копию заключения психолого-медико-педагогической комиссии (для детей с ОВЗ).',reply_markup=markup, parse_mode='Markdown')
    #ПС
    if call.data == 'список':
        markup = types.InlineKeyboardMarkup()
        item1 = types.InlineKeyboardButton(text='Администрация', callback_data='адм')
        item2 = types.InlineKeyboardButton(text='Учителя начальных классов', callback_data='унк')
        item3 = types.InlineKeyboardButton(text='Учителя гуманитарного цикла', callback_data='угц')
        item4 = types.InlineKeyboardButton(text='Учителя физико-математического цикла', callback_data='уфмц')
        item5 = types.InlineKeyboardButton(text='Учителя естественно-научного цикла', callback_data='уенц')
        item6 = types.InlineKeyboardButton(text='Учителя иностранного языка', callback_data='уия')
        item7 = types.InlineKeyboardButton(text='Учителя спортивно-прикладного направления', callback_data='успн')
        item8 = types.InlineKeyboardButton(text='Социально-психологическая служба', callback_data='спс')
        markup.add(item1, item2,item3,item4,item5,item6,item7,item8)
        bot.edit_message_text(text='Выбериите направление', message_id=mesID.id, chat_id=ChatID,reply_markup=markup)
    if call.data == "Нет1":
        global msgTP
        testVivod = True
        bot.delete_message(chat_id=ChatID, message_id=mesID.message_id)
        markup = types.InlineKeyboardMarkup()
        item1 = types.InlineKeyboardButton(text='Да', callback_data='Да2')
        item2 = types.InlineKeyboardButton(text='Нет', callback_data='Нет2')
        markup.add(item1, item2)
        msgTP = bot.send_message(ChatID, 'Обратиться к тех. поддержке?', reply_markup=markup)
    if call.data == "Да2":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(chat_id=ChatID,  text='Ваш вопрос отправлен', reply_markup=markup)
        if testVivod:
            bot.delete_message(chat_id=ChatID, message_id=msgTP.message_id)
    if call.data == "Нет2":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        if testVivod:
            bot.delete_message(chat_id=ChatID, message_id=msgTP.message_id)
        bot.send_message(chat_id=ChatID, text='Ваш вопрос не отправлен', reply_markup=markup)
    if call.data == "Главная":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        bot.send_message(ChatID, "Главная", reply_markup=markup)

    if call.data == "адм" or call.data == "унк" or call.data == "угц" or call.data == "уфмц" or call.data == "уенц" or call.data == "уия" or call.data == "успн" or call.data == "спс":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Задать вопрос")
        item2 = types.KeyboardButton("Сообщить о неполадках")
        item3 = types.KeyboardButton("Предложить идею")
        item4 = types.KeyboardButton("Быстрый доступ")
        markup.add(item1, item2, item3, item4)
        with open(f'Sostavi/{call.data}.txt', encoding='utf8') as file:
            F = file.read()
            bot.send_message(chat_id=ChatID, text=F, reply_markup=markup, parse_mode="Markdown")
    if call.data=="НПИ":
        msg1 = bot.send_message(ChatID, text='Введите имя учителя:')
        bot.register_next_step_handler(msg1, npi)
def npi(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Задать вопрос")
    item2 = types.KeyboardButton("Сообщить о неполадках")
    item3 = types.KeyboardButton("Предложить идею")
    item4 = types.KeyboardButton("Быстрый доступ")
    markup.add(item1, item2, item3, item4)
    maxes = 0
    msgOTV= msgOTVu = ''
    for i in QueryTeachers:
        score = similarity_score(i, message.text)
        if maxes<score:
            maxes=score
            msgOTV=i
    for j in msgOTV:
        if j.isupper():
            msgOTVu+=j
    with open('Sostavi/ОБЩ.txt', encoding="utf8") as file:
        F = file.read()
        print(f'\Sostavi\PHOTOS\{msgOTVu}.jpg')
        img = open(f'Sostavi/PHOTOS/{msgOTVu}.jpg', 'rb')
        bot.send_photo(ChatID, img)
        F=F[F.find(msgOTV):]
        F1 = F[F.find(msgOTV):F.find('■')]
        print(F1)
        bot.send_message(ChatID, F1, reply_markup=markup)

def RaspUCHEN(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Задать вопрос")
    item2 = types.KeyboardButton("Сообщить о неполадках")
    item3 = types.KeyboardButton("Предложить идею")

    item4 = types.KeyboardButton("Быстрый доступ")
    markup.add(item1, item2, item3, item4)
    rowTwo = sheet_ranges1[2]
    rowThree = sheet_ranges1[3]
    colB = colC = None
    for row in sheet_ranges1.iter_cols():
        for cell in row:
            if str(cell.value).lower() == message.text.lower():
                colB = sheet_ranges1[cell.column_letter]
                colC = sheet_ranges1[get_column_letter(cell.col_idx + 1)]
    if colB == None:
        bot.send_message(ChatID, 'Вы ввели класс неправильно или этого класса не существует', reply_markup=markup)
    colA = sheet_ranges1['A']
    mass = ["ПОНЕДЕЛЬНИК:", "ВТОРНИК:", "СРЕДА:", "ЧЕТВЕРГ:", "ПЯТНИЦА:"]
    l=''
    for i in range(len(rowTwo)):
        if message.text == rowTwo[i].value or message.text==rowThree[i].value:
            for j in range(len(colB)):
                if j > 1:
                    if colA[j].value in mass:
                        if l != '':
                            bot.send_message(ChatID,l)
                        l = ''
                        l += colA[j].value + '\n'
                    elif colB[j].value == None or colA[j].value == None:
                        continue
                    else:
                        if colB[j].value != message.text:
                            l += str(colA[j].value) + '. ' + str(colB[j].value) + ' ' + str(colC[j].value) + '\n'
    print(l)
    bot.send_message(ChatID, l, reply_markup=markup)

def RaspUCIT(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Задать вопрос")
    item2 = types.KeyboardButton("Сообщить о неполадках")
    item3 = types.KeyboardButton("Предложить идею")
    item4 = types.KeyboardButton("Быстрый доступ")
    markup.add(item1, item2, item3, item4)

    wb = load_workbook('Raspisanie_UChITELYa_S_9_yanvarya_2024.xlsx')
    sheet_ranges = wb.active
    rowTwo = sheet_ranges[2]
    rowThree = sheet_ranges[3]
    rowNeed = rowNeed1 = rowNeed2 = None
    k = -1
    for row in sheet_ranges['B']:
        k += 1
        if message.text == row.value:
            rowNeed = sheet_ranges[row.row]
            rowNeed1 = sheet_ranges[row.row + 1]
            rowNeed2 = sheet_ranges[row.row + 2]

    mass = ['ПОНЕДЕЛЬНИК:', 'ВТОРНИК:', 'СРЕДА:', 'ЧЕТВЕРГ:', 'ПЯТНИЦА:']
    l = ''
    j = 0
    try:
        for i in range(len(rowNeed)):
            j += 1
            if rowTwo[i].value in mass:
                if l != '':
                    bot.send_message(ChatID, l)
                l = ''
                l += rowTwo[i].value + '\n'
            if message.text == 'Сухая АА' and 69 > j > 2 and rowThree[i].value != None:
                if rowNeed[i].value != None:
                    l += str(rowThree[i].value) + '. ' + str(rowNeed[i].value) + ' физ\n'
                if rowNeed1[i].value != None:
                    l += str(rowThree[i].value) + '. ' + str(rowNeed1[i].value) + ' инф\n'
            elif message.text == 'Шилина МС' and 69 > j > 2 and rowThree[i].value != None:
                if rowNeed[i].value != None:
                    l += str(rowThree[i].value) + '. ' + str(rowNeed[i].value) + ' техн\n'
                if rowNeed1[i].value != None:
                    l += str(rowThree[i].value) + '. ' + str(rowNeed1[i].value) + ' ест\n'
                if rowNeed2[i].value != None:
                    l += str(rowThree[i].value) + '. ' + str(rowNeed2[i].value) + ' хим\n'
            elif rowNeed[i].value != None and 69 > j > 2 and rowThree[i].value != None:
                l += str(rowThree[i].value) + '. ' + str(rowNeed[i].value) + '\n'
        bot.send_message(ChatID, l, reply_markup=markup)
    except:
        bot.send_message(chat_id=ChatID, text='Введите действительное имя или имя в правильном формате (Фамилия ИО)',reply_markup=markup)

#InPoll
bot.infinity_polling()

